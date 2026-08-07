from datetime import date, datetime
from io import BytesIO
import re
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.company_scope import assigned_company_ids
from app.models.case import Case
from app.models.case_visit import CaseVisit
from app.models.master import Bank, Company, CompanyBank, District, Executive, LoanType
from app.models.user import User
from app.schemas.case_import import CaseImportError, CaseImportResponse

HEADERS = ["Visit Type","LOS / Application No","Receive Date","Company","Bank / Finance Company","Applicant","Mobile","Loan Type","Address","District","City","Landmark","Executive","Status","Negative Reason","Remarks"]
REQUIRED = set(HEADERS) - {"Loan Type","Landmark","Negative Reason","Remarks"}
VISIT_TYPES = {"Residence","Office","Permanent","Business","Other"}
STATUSES = {"Pending","Positive","Negative"}

def _key(value): return " ".join(str(value or "").strip().casefold().split())
def _text(value): return " ".join(str(value or "").strip().split()) or None

def template_bytes(db:Session,user:User) -> bytes:
    scope=assigned_company_ids(user)
    company_query=select(Company).where(Company.is_active.is_(True)).order_by(Company.name)
    if scope is not None: company_query=company_query.where(Company.id.in_(scope))
    companies=list(db.scalars(company_query).all()); all_banks=list(db.scalars(select(Bank).order_by(Bank.name)).all())
    inactive={(x.company_id,x.bank_id) for x in db.scalars(select(CompanyBank).where(CompanyBank.is_active.is_(False))).all()}
    executives=list(db.scalars(select(Executive).where(Executive.status=="Active").order_by(Executive.full_name)).all())
    if user.role=="Executive": executives=[user.executive] if user.executive and user.executive.status=="Active" else []
    districts=list(db.scalars(select(District).where(District.is_active.is_(True)).order_by(District.name)).all())
    loans=list(db.scalars(select(LoanType).order_by(LoanType.name)).all())
    wb=Workbook();ws=wb.active;ws.title="Case Import";ws.append(HEADERS);ws.freeze_panes="A2";ws.auto_filter.ref="A1:P500";ws.sheet_view.zoomScale=90
    required_fill=PatternFill("solid",fgColor="FFFBEA");optional_fill=PatternFill("solid",fgColor="FFFFFF")
    header_fill=PatternFill("solid",fgColor="0F172A");thin=Side(style="thin",color="D7DEE8");cell_border=Border(left=thin,right=thin,top=thin,bottom=thin)
    widths=[14,22,15,28,28,24,16,18,42,18,18,24,26,14,28,32];left_aligned={2,4,5,6,9,12,13,15,16};ws.row_dimensions[1].height=34
    for i,(name,width) in enumerate(zip(HEADERS,widths),1):
        cell=ws.cell(1,i);cell.font=Font(bold=True,color="FFFFFF");cell.fill=header_fill;cell.border=cell_border;cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);ws.column_dimensions[get_column_letter(i)].width=width
        for row in range(2,501):
            data_cell=ws.cell(row,i);data_cell.fill=required_fill if name in REQUIRED else optional_fill;data_cell.border=cell_border;data_cell.alignment=Alignment(horizontal="left" if i in left_aligned else "center",vertical="center",wrap_text=i in {9,15,16})
    for row in range(2,501):ws.row_dimensions[row].height=23;ws.cell(row,3).number_format="yyyy-mm-dd";ws.cell(row,2).number_format="@";ws.cell(row,7).number_format="@"
    company_sheet=wb.create_sheet("Companies");company_sheet.append(["Company","Bank Range"])
    bank_sheet=wb.create_sheet("Company Banks");bank_sheet.append(["Company","Bank / Finance Company"])
    for company in companies:
        range_name=f"CompanyBanks_{company.id}";company_sheet.append([company.name,range_name]);start=bank_sheet.max_row+1
        valid=[bank for bank in all_banks if (company.id,bank.id) not in inactive]
        for bank in valid:bank_sheet.append([company.name,bank.name])
        end=bank_sheet.max_row
        if end>=start:wb.defined_names.add(DefinedName(range_name,attr_text=f"'Company Banks'!$B${start}:$B${end}"))
    exec_sheet=wb.create_sheet("Executives");exec_sheet.append(["Executive"])
    for x in executives:exec_sheet.append([x.full_name])
    district_sheet=wb.create_sheet("Districts");district_sheet.append(["District"])
    for x in districts:district_sheet.append([x.name])
    loan_sheet=wb.create_sheet("Loan Types");loan_sheet.append(["Loan Type"])
    for x in loans:loan_sheet.append([x.name])
    instructions=wb.create_sheet("Instructions");instructions.sheet_view.zoomScale=90;instructions.freeze_panes="A3";instructions.merge_cells("A1:D1")
    title=instructions["A1"];title.value="Case Import Instructions";title.font=Font(bold=True,color="FFFFFF",size=16);title.fill=header_fill;title.alignment=Alignment(vertical="center");instructions.row_dimensions[1].height=32
    instructions.append(["Field / Topic","Requirement","Example","Notes"])
    instruction_rows=[("Required vs optional","Yellow cells are required; white cells are optional.","","Complete every required field before upload."),("Headers","Do not rename, remove, reorder, or add headers.","","The importer requires the supplied column layout."),("Visit Type","Accepted values: Residence, Office, Permanent, Business, Other.","Residence","Use the dropdown."),("Status","Accepted values: Pending, Positive, Negative.","Pending","Use the dropdown."),("Receive Date","Use yyyy-mm-dd.","2026-08-07","Dates are validated during import."),("Company -> Bank","Select Company first, then select its Bank / Finance Company.","Acme Finance -> Example Bank","The Bank dropdown depends on the Company in the same row."),("Dropdowns","Choose exact active master values.","","Backend validation remains authoritative."),("Formulas","Do not paste formulas into the import sheet.","","Paste values only."),("Negative Reason","Required only when Status is Negative.","Applicant unavailable","Optional for other statuses.")]
    for row in instruction_rows:instructions.append(row)
    instructions.append([]);instructions.append(["Example Import Row (copy values only; do not copy formulas)"]);instructions.append(HEADERS);instructions.append(["Residence","LOS-0001","2026-08-07","Example Company","Example Bank","Example Applicant","0987654321","Home Loan","Plot 12, Main Road","Jaipur","Jaipur","Near Central Park","Example Executive","Pending","","Initial verification visit"])
    for column,width in zip("ABCD",(24,52,28,42)):instructions.column_dimensions[column].width=width
    for cell in instructions[2]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=header_fill;cell.border=cell_border;cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    example_header_row=len(instruction_rows)+5
    for cell in instructions[example_header_row]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=header_fill;cell.border=cell_border;cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for row in instructions.iter_rows(min_row=3,max_row=instructions.max_row):
        for cell in row:cell.alignment=Alignment(vertical="top",wrap_text=True)
    if companies:
        wb.defined_names.add(DefinedName("CompanyList",attr_text=f"Companies!$A$2:$A${company_sheet.max_row}"));wb.defined_names.add(DefinedName("CompanyRangeLookup",attr_text=f"Companies!$A$2:$B${company_sheet.max_row}"))
    for name,sheet in (("ExecutiveList",exec_sheet),("DistrictList",district_sheet),("LoanTypeList",loan_sheet)):
        if sheet.max_row>=2:wb.defined_names.add(DefinedName(name,attr_text=f"'{sheet.title}'!$A$2:$A${sheet.max_row}"))
    validations=((1,'"Residence,Office,Permanent,Business,Other"'),(4,"=CompanyList"),(5,'=INDIRECT(VLOOKUP($D2,CompanyRangeLookup,2,FALSE))'),(8,"=LoanTypeList"),(10,"=DistrictList"),(13,"=ExecutiveList"),(14,'"Pending,Positive,Negative"'))
    for column,formula in validations:
        dv=DataValidation(type="list",formula1=formula,allow_blank=HEADERS[column-1] not in REQUIRED);dv.error="Select a value from the dropdown.";dv.errorTitle="Invalid value";dv.showErrorMessage=True;dv.errorStyle="stop";ws.add_data_validation(dv);dv.add(f"{get_column_letter(column)}2:{get_column_letter(column)}500")
    for sheet in (company_sheet,bank_sheet,exec_sheet,district_sheet,loan_sheet):sheet.sheet_state="hidden"
    out=BytesIO();wb.save(out);return out.getvalue()

def _date(value):
    if isinstance(value,datetime): return value.date()
    if isinstance(value,date): return value
    text=str(value or "").strip()
    for fmt in ("%Y-%m-%d","%d-%m-%Y","%d/%m/%Y"):
        try:return datetime.strptime(text,fmt).date()
        except ValueError:pass
    return None

def import_cases(db:Session,user:User,content:bytes)->CaseImportResponse:
    errors=[]
    try: wb=load_workbook(BytesIO(content),read_only=True,data_only=False)
    except Exception: return CaseImportResponse(success=False,errors=[CaseImportError(row=0,field="File",message="Invalid XLSX file")])
    if "Case Import" not in wb.sheetnames:return CaseImportResponse(success=False,errors=[CaseImportError(row=0,field="Sheet",message="Case Import sheet not found")])
    ws=wb["Case Import"]; supplied=[_text(x.value) for x in next(ws.iter_rows(min_row=1,max_row=1),[])]
    if supplied!=HEADERS:return CaseImportResponse(success=False,errors=[CaseImportError(row=1,field="Columns",message="Template columns do not match")])
    companies={_key(x.name):x for x in db.scalars(select(Company)).all()};banks={_key(x.name):x for x in db.scalars(select(Bank)).all()};districts={_key(x.name):x for x in db.scalars(select(District)).all()};executives={_key(x.full_name):x for x in db.scalars(select(Executive)).all()}
    scope=assigned_company_ids(user); parsed=[]; seen=set()
    for number,cells in enumerate(ws.iter_rows(min_row=2),2):
        values={header:cell.value for header,cell in zip(HEADERS,cells)}
        if all(v is None or str(v).strip()=="" for v in values.values()):continue
        for field in REQUIRED:
            if _text(values[field]) is None:errors.append(CaseImportError(row=number,field=field,value=None,message="Required"))
        for field,value in values.items():
            if isinstance(value,str) and value.startswith("="):errors.append(CaseImportError(row=number,field=field,value=value,message="Formulas are not allowed"))
        company=companies.get(_key(values["Company"]));bank=banks.get(_key(values["Bank / Finance Company"]));district=districts.get(_key(values["District"]));executive=executives.get(_key(values["Executive"]));received=_date(values["Receive Date"])
        if not company or not company.is_active:errors.append(CaseImportError(row=number,field="Company",value=_text(values["Company"]),message="Active company not found"))
        elif scope is not None and company.id not in scope:errors.append(CaseImportError(row=number,field="Company",value=company.name,message="Company is not assigned to this user"))
        if not bank:errors.append(CaseImportError(row=number,field="Bank / Finance Company",value=_text(values["Bank / Finance Company"]),message="Bank not found"))
        if company and bank:
            mapping=db.scalar(select(CompanyBank).where(CompanyBank.company_id==company.id,CompanyBank.bank_id==bank.id))
            if mapping is not None and not mapping.is_active:errors.append(CaseImportError(row=number,field="Bank / Finance Company",value=bank.name,message="Company-bank association is inactive"))
        if not district or not district.is_active:errors.append(CaseImportError(row=number,field="District",value=_text(values["District"]),message="Active district not found"))
        if not executive or executive.status!="Active":errors.append(CaseImportError(row=number,field="Executive",value=_text(values["Executive"]),message="Active Executive not found"))
        elif user.role=="Executive" and (not user.executive or user.executive.id!=executive.id):errors.append(CaseImportError(row=number,field="Executive",value=executive.full_name,message="Executives may import only their own visits"))
        if _text(values["Visit Type"]) not in VISIT_TYPES:errors.append(CaseImportError(row=number,field="Visit Type",value=_text(values["Visit Type"]),message="Invalid visit type"))
        status=_text(values["Status"])
        if status not in STATUSES:errors.append(CaseImportError(row=number,field="Status",value=status,message="Invalid status"))
        if status=="Negative" and not _text(values["Negative Reason"]):errors.append(CaseImportError(row=number,field="Negative Reason",message="Required when Status is Negative"))
        if not received:errors.append(CaseImportError(row=number,field="Receive Date",value=_text(values["Receive Date"]),message="Invalid date; use YYYY-MM-DD"))
        mobile=re.sub(r"\D","",str(values["Mobile"] or ""))
        if not 10<=len(mobile)<=15:errors.append(CaseImportError(row=number,field="Mobile",value=_text(values["Mobile"]),message="Mobile must contain 10 to 15 digits"))
        identity=(_key(values["Company"]),_key(values["Bank / Finance Company"]),_key(values["LOS / Application No"]),_key(values["Visit Type"]),received)
        if identity in seen:errors.append(CaseImportError(row=number,field="Row",message="Duplicate visit row in file"))
        if company and bank and _text(values["LOS / Application No"]):
            current=db.scalar(select(Case).where(Case.company_id==company.id,
                func.lower(func.trim(Case.bank))==_key(bank.name),
                func.lower(func.trim(Case.los_no))==_key(values["LOS / Application No"])))
            if current is not None and _key(current.applicant)!=_key(values["Applicant"]):
                errors.append(CaseImportError(row=number,field="Applicant",value=_text(values["Applicant"]),message="Existing application has a different applicant"))
        seen.add(identity);parsed.append((number,values,company,bank,district,executive,received,mobile))
    if errors:return CaseImportResponse(success=False,errors=errors)
    created,visits,existing=0,0,set()
    try:
        for _,v,company,bank,district,executive,received,mobile in parsed:
            los=_text(v["LOS / Application No"])
            parent=db.scalar(select(Case).where(Case.company_id==company.id,func.lower(func.trim(Case.bank))==_key(bank.name),func.lower(func.trim(Case.los_no))==_key(los)).with_for_update())
            if parent is None:
                parent=Case(case_no=f"JA-{uuid4().hex[:12].upper()}",los_no=los,receive_date=received,company_id=company.id,company=company.name,bank=bank.name,applicant=_text(v["Applicant"]),mobile=mobile,loan_type=_text(v["Loan Type"]),address=_text(v["Address"]),district_id=district.id,district=district.name,city=_text(v["City"]),executive=executive.full_name,status=_text(v["Status"]));db.add(parent);db.flush();created+=1
            else: existing.add(parent.id)
            status=_text(v["Status"]);db.add(CaseVisit(case_id=parent.id,visit_type=_text(v["Visit Type"]),address=_text(v["Address"]),district_id=district.id,district=district.name,city=_text(v["City"]),landmark=_text(v["Landmark"]),executive=executive.full_name,status=status,negative_reason=_text(v["Negative Reason"]),receive_date=received,closed_date=date.today() if status in {"Positive","Negative"} else None,remarks=_text(v["Remarks"]),created_by_user_id=user.id,updated_by_user_id=user.id));visits+=1
        db.commit()
    except Exception:db.rollback();raise
    return CaseImportResponse(success=True,created_applications=created,created_visits=visits,updated_existing_applications=len(existing))
