from datetime import date, datetime, timedelta, timezone
from difflib import SequenceMatcher
from io import BytesIO
import json
import re
import secrets

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.company_scope import assigned_company_ids
from app.models.case import Case
from app.models.case_import import CaseImportRow, CaseImportSession
from app.models.case_visit import CaseVisit
from app.models.master import Bank, Company, CompanyBank, District, Executive, LoanType
from app.models.user import User
from app.schemas.case_import import ImportCommitResponse, ImportPreviewResponse
from app.services.case_import_service import HEADERS, REQUIRED, STATUSES, VISIT_TYPES, _date, _key, _text

FIELDS = {
    "Visit Type":"visit_type","LOS / Application No":"los_no","Receive Date":"receive_date","Company":"company",
    "Bank / Finance Company":"bank","Applicant":"applicant","Mobile":"mobile","Loan Type":"loan_type",
    "Address":"address","District":"district","City":"city","Landmark":"landmark","Executive":"executive",
    "Status":"status","Negative Reason":"negative_reason","Remarks":"remarks",
}
REVERSE_FIELDS = {value:key for key,value in FIELDS.items()}


def _header_key(value):
    return re.sub(r"[^a-z0-9]+","",str(value or "").casefold())


HEADER_KEYS = {_header_key(header):header for header in HEADERS}


def _clean(value):
    if isinstance(value,(date,datetime)): return value.isoformat()[:10]
    return _text(value) or ""


def _punctuation_key(value):
    return re.sub(r"[^a-z0-9]+","",str(value or "").casefold())


def _issue(field,value,message,suggestion=None):
    result={"field":field,"entered_value":_text(value),"message":message}
    if suggestion:
        result.update(suggested_value=suggestion[1],suggested_id=suggestion[0],confidence="high")
    return result


def _match(value,items,name_attr="name"):
    entered=_text(value)
    if not entered:return None,None
    exact=next((item for item in items if getattr(item,name_attr)==entered),None)
    if exact:return exact,None
    folded=next((item for item in items if getattr(item,name_attr).casefold()==entered.casefold()),None)
    if folded:return folded,None
    compact=_punctuation_key(entered)
    normalized=next((item for item in items if _punctuation_key(getattr(item,name_attr))==compact),None)
    if normalized:return normalized,None
    scored=sorted(((SequenceMatcher(None,compact,_punctuation_key(getattr(item,name_attr))).ratio(),item) for item in items),reverse=True,key=lambda pair:pair[0])
    if scored and scored[0][0]>=0.72:
        item=scored[0][1];return None,(item.id,getattr(item,name_attr))
    return None,None


def _masters(db,user):
    scope=assigned_company_ids(user)
    cq=select(Company).where(Company.is_active.is_(True)).order_by(Company.name)
    if scope is not None:cq=cq.where(Company.id.in_(scope))
    companies=list(db.scalars(cq).all())
    banks=list(db.scalars(select(Bank).order_by(Bank.name)).all())
    inactive={(x.company_id,x.bank_id) for x in db.scalars(select(CompanyBank).where(CompanyBank.is_active.is_(False))).all()}
    executives=list(db.scalars(select(Executive).where(Executive.status=="Active").order_by(Executive.full_name)).all())
    if user.role=="Executive":executives=[user.executive] if user.executive and user.executive.status=="Active" else []
    districts=list(db.scalars(select(District).where(District.is_active.is_(True)).order_by(District.name)).all())
    loans=list(db.scalars(select(LoanType).order_by(LoanType.name)).all())
    return companies,banks,inactive,executives,districts,loans


def _validate(db,user,data,duplicate_row=None):
    companies,banks,inactive,executives,districts,loans=_masters(db,user)
    errors=[];warnings=[];resolved=dict(data)
    company,company_suggestion=_match(data.get("company"),companies)
    if not company:errors.append(_issue("company",data.get("company"),"Active company not found or not available in your scope",company_suggestion))
    else:resolved["company"]=company.name
    available_banks=[bank for bank in banks if not company or (company.id,bank.id) not in inactive]
    bank,bank_suggestion=_match(data.get("bank"),available_banks)
    if not bank:errors.append(_issue("bank",data.get("bank"),"Bank is not valid for the selected company",bank_suggestion))
    else:resolved["bank"]=bank.name
    executive,exec_suggestion=_match(data.get("executive"),executives,"full_name")
    if not executive:errors.append(_issue("executive",data.get("executive"),"Active Executive not found or not permitted",exec_suggestion))
    else:resolved["executive"]=executive.full_name
    district,district_suggestion=_match(data.get("district"),districts)
    if not district:errors.append(_issue("district",data.get("district"),"Active district not found",district_suggestion))
    else:resolved["district"]=district.name
    if _text(data.get("loan_type")):
        loan,loan_suggestion=_match(data.get("loan_type"),loans)
        if not loan:errors.append(_issue("loan_type",data.get("loan_type"),"Loan type not found",loan_suggestion))
        else:resolved["loan_type"]=loan.name
    for field in (FIELDS[x] for x in REQUIRED):
        if not _text(data.get(field)):errors.append(_issue(field,data.get(field),"Required"))
    received=_date(data.get("receive_date"))
    if not received:errors.append(_issue("receive_date",data.get("receive_date"),"Invalid date; use YYYY-MM-DD"))
    else:resolved["receive_date"]=received.isoformat()
    if _text(data.get("visit_type")) not in VISIT_TYPES:errors.append(_issue("visit_type",data.get("visit_type"),"Invalid visit type"))
    if _text(data.get("status")) not in STATUSES:errors.append(_issue("status",data.get("status"),"Invalid status"))
    if _text(data.get("status"))=="Negative" and not _text(data.get("negative_reason")):errors.append(_issue("negative_reason",data.get("negative_reason"),"Required when Status is Negative"))
    mobile=re.sub(r"\\D","",str(data.get("mobile") or ""));resolved["mobile"]=mobile
    if not 10<=len(mobile)<=15:errors.append(_issue("mobile",data.get("mobile"),"Mobile must contain 10 to 15 digits"))
    if duplicate_row:errors.append(_issue("row",None,f"Duplicate of Excel row {duplicate_row}"))
    existing=None
    if company and bank and _text(data.get("los_no")):
        existing=db.scalar(select(Case).where(Case.company_id==company.id,func.lower(func.trim(Case.bank))==_key(bank.name),func.lower(func.trim(Case.los_no))==_key(data.get("los_no"))))
        if existing and _key(existing.applicant)!=_key(data.get("applicant")):errors.append(_issue("applicant",data.get("applicant"),"Existing application has a different applicant"))
    if mobile and 10<=len(mobile)<=15:
        count=db.scalar(select(func.count(CaseVisit.id)).join(Case).where(Case.mobile==mobile)) or 0
        if count:warnings.append(_issue("mobile",mobile,f"Mobile already exists in {count} visits."))
    action="ADD_VISIT_TO_EXISTING_APPLICATION" if existing else "CREATE_APPLICATION_AND_VISIT"
    state="ERROR" if errors else ("WARNING" if warnings else "VALID")
    return resolved,state,action,errors,warnings


def _summary(rows):
    states=[row.state for row in rows]
    return {"total_rows":len(rows),"valid_rows":states.count("VALID"),"warning_rows":states.count("WARNING"),"error_rows":states.count("ERROR"),"imported_rows":states.count("IMPORTED"),"new_applications":sum(row.intended_action=="CREATE_APPLICATION_AND_VISIT" and row.state!="IMPORTED" for row in rows),"new_visits_existing_application":sum(row.intended_action=="ADD_VISIT_TO_EXISTING_APPLICATION" and row.state!="IMPORTED" for row in rows)}


def _options(db,user):
    companies,banks,inactive,executives,districts,loans=_masters(db,user)
    return {"companies":[{"id":x.id,"name":x.name} for x in companies],"banks_by_company":{str(c.id):[{"id":b.id,"name":b.name} for b in banks if (c.id,b.id) not in inactive] for c in companies},"executives":[{"id":x.id,"name":x.full_name} for x in executives],"districts":[{"id":x.id,"name":x.name} for x in districts],"loan_types":[{"id":x.id,"name":x.name} for x in loans],"visit_types":sorted(VISIT_TYPES),"statuses":sorted(STATUSES)}


def _response(db,user,session):
    rows=list(session.rows)
    return ImportPreviewResponse(import_token=session.token,filename=session.filename,uploaded_at=session.created_at.isoformat(),expires_at=session.expires_at.isoformat(),summary=_summary(rows),rows=[{"row_number":r.row_number,"data":json.loads(r.data_json),"state":r.state,"intended_action":r.intended_action,"errors":json.loads(r.errors_json),"warnings":json.loads(r.warnings_json)} for r in rows],options=_options(db,user))


def preview_import(db:Session,user:User,content:bytes,filename:str):
    wb=load_workbook(BytesIO(content),read_only=True,data_only=False)
    if "Case Import" not in wb.sheetnames:raise ValueError("Case Import sheet not found")
    ws=wb["Case Import"];raw_headers=[cell.value for cell in next(ws.iter_rows(min_row=1,max_row=1),[])]
    canonical=[];unknown=[]
    for header in raw_headers:
        match=HEADER_KEYS.get(_header_key(header))
        if not match:unknown.append(str(header or "blank"))
        canonical.append(match)
    if unknown or len(canonical)!=len(HEADERS) or set(canonical)!=set(HEADERS):raise ValueError("Template columns do not match. Unknown or ambiguous columns: "+", ".join(unknown))
    now=datetime.now(timezone.utc);session=CaseImportSession(token=secrets.token_urlsafe(32),user_id=user.id,filename=filename[:255],created_at=now,expires_at=now+timedelta(hours=2));db.add(session);db.flush()
    seen={}
    for number,cells in enumerate(ws.iter_rows(min_row=2),2):
        raw={FIELDS[header]:_clean(cell.value) for header,cell in zip(canonical,cells)}
        if not any(raw.values()):continue
        signature=tuple(_punctuation_key(raw.get(FIELDS[h])) for h in HEADERS)
        duplicate=seen.get(signature);seen.setdefault(signature,number)
        data,state,action,errors,warnings=_validate(db,user,raw,duplicate)
        session.rows.append(CaseImportRow(row_number=number,data_json=json.dumps(data),state=state,intended_action=action,errors_json=json.dumps(errors),warnings_json=json.dumps(warnings)))
    db.commit();db.refresh(session);return _response(db,user,session)


def get_session(db,user,token=None):
    query=select(CaseImportSession).where(CaseImportSession.user_id==user.id,CaseImportSession.expires_at>datetime.now(timezone.utc))
    query=query.where(CaseImportSession.token==token) if token else query.order_by(CaseImportSession.created_at.desc())
    session=db.scalar(query)
    if not session:raise PermissionError("Import session not found or expired")
    return session


def resume_import(db,user):
    return _response(db,user,get_session(db,user))


def _create_visit(db,user,data):
    companies,banks,_,executives,districts,_=_masters(db,user)
    company=next(x for x in companies if x.name==data["company"]);bank=next(x for x in banks if x.name==data["bank"]);district=next(x for x in districts if x.name==data["district"]);executive=next(x for x in executives if x.full_name==data["executive"])
    parent=db.scalar(select(Case).where(Case.company_id==company.id,func.lower(func.trim(Case.bank))==_key(bank.name),func.lower(func.trim(Case.los_no))==_key(data["los_no"])).with_for_update())
    created=False
    if parent is None:
        from uuid import uuid4
        parent=Case(case_no=f"JA-{uuid4().hex[:12].upper()}",los_no=_text(data["los_no"]),receive_date=_date(data["receive_date"]),company_id=company.id,company=company.name,bank=bank.name,applicant=_text(data["applicant"]),mobile=data["mobile"],loan_type=_text(data.get("loan_type")),address=_text(data.get("address")),district_id=district.id,district=district.name,city=_text(data.get("city")),executive=executive.full_name,status=data["status"]);db.add(parent);db.flush();created=True
    status=data["status"];db.add(CaseVisit(case_id=parent.id,visit_type=data["visit_type"],address=_text(data.get("address")),district_id=district.id,district=district.name,city=_text(data.get("city")),landmark=_text(data.get("landmark")),executive=executive.full_name,status=status,negative_reason=_text(data.get("negative_reason")),receive_date=_date(data["receive_date"]),closed_date=date.today() if status in {"Positive","Negative"} else None,remarks=_text(data.get("remarks")),created_by_user_id=user.id,updated_by_user_id=user.id))
    return created


def commit_import(db:Session,user:User,token:str,selections):
    session=get_session(db,user,token);by_number={row.row_number:row for row in session.rows}
    imported=created=added=0;failed=[]
    for selection in selections:
        row=by_number.get(selection.row_number)
        if not row:failed.append({"row_number":selection.row_number,"errors":[_issue("row",None,"Row is not part of this import session")]});continue
        if row.imported_at:failed.append({"row_number":row.row_number,"errors":[_issue("row",None,"Row was already imported")]});continue
        data,state,action,errors,warnings=_validate(db,user,selection.resolved_data)
        row.data_json=json.dumps(data);row.state=state;row.intended_action=action;row.errors_json=json.dumps(errors);row.warnings_json=json.dumps(warnings)
        if state=="ERROR":db.commit();failed.append({"row_number":row.row_number,"errors":errors});continue
        try:
            was_created=_create_visit(db,user,data);row.state="IMPORTED";row.imported_at=datetime.now(timezone.utc);db.commit();imported+=1;created+=int(was_created);added+=int(not was_created)
        except Exception as exc:
            db.rollback();failed.append({"row_number":row.row_number,"errors":[_issue("row",None,f"Import failed: {str(exc)[:160]}")]})
    db.refresh(session);remaining=sum(row.state!="IMPORTED" for row in session.rows)
    if not remaining:session.completed_at=datetime.now(timezone.utc);db.commit()
    return ImportCommitResponse(success=not failed,imported_rows=imported,created_applications=created,created_visits=imported,added_to_existing_applications=added,failed_rows=failed,remaining_rows=remaining)
