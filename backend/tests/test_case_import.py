import os,unittest
from datetime import date
from io import BytesIO
os.environ.setdefault("DATABASE_URL","sqlite://")
from openpyxl import Workbook,load_workbook
from sqlalchemy import create_engine,select
from sqlalchemy.orm import Session
import app.db.base  # noqa
from app.db.database import Base
from app.models.case import Case
from app.models.case_visit import CaseVisit
from app.models.master import Bank,Company,CompanyBank,District,Executive,LoanType
from app.models.user import User,UserCompany
from app.services.case_import_service import HEADERS,import_cases,template_bytes
from app.services.smart_case_import_service import commit_import,preview_import,resume_import
from app.schemas.case_import import ImportCommitRow
from app.services.dashboard_service import get_dashboard_summary
from app.services.monthly_billing_service import monthly_billing

class CaseImportTests(unittest.TestCase):
 def setUp(self):
  self.engine=create_engine("sqlite://");Base.metadata.create_all(self.engine);self.db=Session(self.engine)
  self.company=Company(name="Agency",is_active=True);self.other=Company(name="Other",is_active=True);self.inactive_company=Company(name="Inactive Co",is_active=False);self.bank=Bank(name="AU Bank");self.bank2=Bank(name="BOB");self.district=District(name="Jaipur",state="Rajasthan",is_active=True);self.inactive_district=District(name="Inactive District",state="Rajasthan",is_active=False);self.executive=Executive(full_name="Exec One",status="Active");self.inactive_executive=Executive(full_name="Inactive Exec",status="Inactive");self.loan=LoanType(name="Home Loan");self.admin=User(full_name="Admin",username="import-admin",email="import-admin@test",password_hash="x",role="Admin");self.db.add_all([self.company,self.other,self.inactive_company,self.bank,self.bank2,self.district,self.inactive_district,self.executive,self.inactive_executive,self.loan,self.admin]);self.db.commit()
 def tearDown(self):self.db.close();self.engine.dispose()
 def book(self,rows):
  wb=Workbook();ws=wb.active;ws.title="Case Import";ws.append(HEADERS)
  for row in rows:ws.append(row)
  out=BytesIO();wb.save(out);return out.getvalue()
 def row(self,**changes):
  values={"Visit Type":"Residence","LOS / Application No":"LOS-1","Receive Date":date(2026,8,1),"Company":"Agency","Bank / Finance Company":"AU Bank","Applicant":"Applicant","Mobile":"9876543210","Loan Type":"","Address":"Address","District":"Jaipur","City":"Jaipur","Landmark":"","Executive":"Exec One","Status":"Pending","Negative Reason":"","Remarks":""};values.update(changes);return [values[x] for x in HEADERS]
 def test_template_and_valid_multi_visit_identity(self):
  wb=load_workbook(BytesIO(template_bytes(self.db,self.admin)));self.assertEqual(wb.sheetnames,["Case Import","Companies","Company Banks","Executives","Districts","Loan Types","Instructions"]);self.assertEqual([x.value for x in wb["Case Import"][1]],HEADERS)
  self.assertEqual(wb["Case Import"].freeze_panes,"A2");self.assertEqual(wb["Case Import"]["B2"].number_format,"@");self.assertEqual(wb["Case Import"]["G2"].number_format,"@")
  sheet=wb["Case Import"];self.assertEqual(sheet.auto_filter.ref,"A1:P500");self.assertEqual(sheet.sheet_view.zoomScale,90);self.assertEqual(sheet.row_dimensions[1].height,34);self.assertEqual(sheet.row_dimensions[500].height,23)
  self.assertEqual([sheet.column_dimensions[x].width for x in "ABCDEFGHIJKLMNOP"],[14,22,15,28,28,24,16,18,42,18,18,24,26,14,28,32])
  self.assertEqual(sheet["A1"].fill.fgColor.rgb,"000F172A");self.assertEqual(sheet["A1"].font.color.rgb,"00FFFFFF");self.assertTrue(sheet["A1"].alignment.wrap_text);self.assertEqual(sheet["A1"].border.left.style,"thin")
  self.assertEqual(sheet["A2"].fill.fgColor.rgb,"00FFFBEA");self.assertEqual(sheet["H2"].fill.fgColor.rgb,"00FFFFFF");self.assertEqual(sheet["B2"].alignment.horizontal,"left");self.assertEqual(sheet["A2"].alignment.horizontal,"center")
  self.assertIsNone(sheet["A2"].value);self.assertEqual(wb["Instructions"]["A1"].value,"Case Import Instructions");self.assertTrue(any(cell.value=="Residence" for row in wb["Instructions"].iter_rows() for cell in row))
  self.assertIn("Agency",[x[0].value for x in wb["Companies"].iter_rows(min_row=2)]);self.assertNotIn("Inactive Co",[x[0].value for x in wb["Companies"].iter_rows(min_row=2)])
  self.assertIn("Exec One",[x[0].value for x in wb["Executives"].iter_rows(min_row=2)]);self.assertNotIn("Inactive Exec",[x[0].value for x in wb["Executives"].iter_rows(min_row=2)])
  self.assertIn("Jaipur",[x[0].value for x in wb["Districts"].iter_rows(min_row=2)]);self.assertIn("Home Loan",[x[0].value for x in wb["Loan Types"].iter_rows(min_row=2)])
  self.assertNotIn("Inactive District",[x[0].value for x in wb["Districts"].iter_rows(min_row=2)]);self.assertIn(("Agency","AU Bank"),[(x[0].value,x[1].value) for x in wb["Company Banks"].iter_rows(min_row=2)])
  validations=wb["Case Import"].data_validations.dataValidation;formulas={dv.formula1 for dv in validations};self.assertIn('"Residence,Office,Permanent,Business,Other"',formulas);self.assertIn('"Pending,Positive,Negative"',formulas);self.assertTrue(any("VLOOKUP" in x for x in formulas));self.assertTrue(all("500" in str(dv.sqref) for dv in validations))
  result=import_cases(self.db,self.admin,self.book([self.row(),self.row(**{"Visit Type":"Office","Receive Date":date(2026,8,2)})]));self.assertTrue(result.success);self.assertEqual((result.created_applications,result.created_visits),(1,2));self.assertEqual(self.db.query(CaseVisit).count(),2)
 def test_different_bank_creates_parent_and_existing_adds_visit(self):
  first=import_cases(self.db,self.admin,self.book([self.row()]));self.assertTrue(first.success)
  result=import_cases(self.db,self.admin,self.book([self.row(**{"Visit Type":"Office"}),self.row(**{"Bank / Finance Company":"BOB"})]));self.assertTrue(result.success);self.assertEqual((result.created_applications,result.updated_existing_applications),(1,1));self.assertEqual(self.db.query(Case).count(),2)
 def test_one_invalid_row_rolls_back_and_duplicate_rejected(self):
  result=import_cases(self.db,self.admin,self.book([self.row(),self.row(**{"Company":"Missing"})]));self.assertFalse(result.success);self.assertEqual(self.db.query(Case).count(),0)
  duplicate=import_cases(self.db,self.admin,self.book([self.row(),self.row()]));self.assertFalse(duplicate.success);self.assertTrue(any(x.message=="Duplicate visit row in file" for x in duplicate.errors))
 def test_negative_reason_and_manager_scope(self):
  manager=User(full_name="Manager",username="import-manager",email="import-manager@test",password_hash="x",role="Manager");self.db.add(manager);self.db.flush();self.db.add(UserCompany(user_id=manager.id,company_id=self.company.id));self.db.commit()
  negative=import_cases(self.db,manager,self.book([self.row(**{"Status":"Negative"})]));self.assertFalse(negative.success)
  scoped=import_cases(self.db,manager,self.book([self.row(**{"Company":"Other"})]));self.assertFalse(scoped.success);self.assertTrue(any("not assigned" in x.message for x in scoped.errors))
  wb=load_workbook(BytesIO(template_bytes(self.db,manager)));self.assertEqual([x[0].value for x in wb["Companies"].iter_rows(min_row=2)],["Agency"])
 def test_invalid_masters_and_executive_own_scope(self):
  for field,value in (("Bank / Finance Company","Missing"),("District","Missing"),("Executive","Missing")):
   result=import_cases(self.db,self.admin,self.book([self.row(**{field:value})]));self.assertFalse(result.success);self.assertEqual(self.db.query(Case).count(),0)
  other_exec=Executive(full_name="Other Exec",status="Active");executive_user=User(full_name="Exec User",username="import-exec",email="import-exec@test",password_hash="x",role="Executive",executive=self.executive);self.db.add_all([other_exec,executive_user]);self.db.flush();self.db.add(UserCompany(user_id=executive_user.id,company_id=self.company.id));self.db.commit()
  result=import_cases(self.db,executive_user,self.book([self.row(**{"Executive":"Other Exec"})]));self.assertFalse(result.success);self.assertTrue(any("only their own" in x.message for x in result.errors))
 def test_imported_visit_is_visible_to_dashboard_and_billing(self):
  result=import_cases(self.db,self.admin,self.book([self.row(**{"Status":"Positive"})]));self.assertTrue(result.success)
  dashboard=get_dashboard_summary(self.db);self.assertEqual((dashboard.total_cases,dashboard.positive_cases),(1,1))
  billing=monthly_billing(self.db,"2026-08");self.assertEqual(billing.summary.billable_cases,1)

 def test_smart_preview_suggestion_duplicate_and_partial_commit(self):
  wb=Workbook();ws=wb.active;ws.title="Case Import";ws.append(["visit type","LOS Application No","receive date","company","bank / finance company","applicant","mobile","loan type","address","district","city","landmark","executive","status","negative reason","remarks"])
  ws.append(self.row(**{"Bank / Finance Company":"AU Bnak"}));ws.append(self.row(**{"Bank / Finance Company":"AU Bnak"}));out=BytesIO();wb.save(out)
  preview=preview_import(self.db,self.admin,out.getvalue(),"smart.xlsx");self.assertEqual((preview.summary.total_rows,preview.summary.error_rows),(2,2))
  bank_error=next(x for x in preview.rows[0].errors if x.field=="bank");self.assertEqual((bank_error.suggested_value,bank_error.confidence),("AU Bank","high"));self.assertTrue(any("Duplicate of Excel row 2" in x.message for x in preview.rows[1].errors))
  fixed=dict(preview.rows[0].data);fixed["bank"]="AU Bank";result=commit_import(self.db,self.admin,preview.import_token,[ImportCommitRow(row_number=2,resolved_data=fixed)])
  self.assertEqual((result.imported_rows,result.created_applications,result.remaining_rows),(1,1,1));self.assertEqual(self.db.query(Case).count(),1)
  repeated=commit_import(self.db,self.admin,preview.import_token,[ImportCommitRow(row_number=2,resolved_data=fixed)]);self.assertIn("already imported",repeated.failed_rows[0].errors[0].message)

 def test_smart_warning_resume_ownership_and_add_visit(self):
  self.assertTrue(import_cases(self.db,self.admin,self.book([self.row()])).success)
  preview=preview_import(self.db,self.admin,self.book([self.row(**{"Visit Type":"Office","Receive Date":date(2026,8,2)})]),"resume.xlsx")
  self.assertEqual((preview.rows[0].state,preview.rows[0].intended_action),("WARNING","ADD_VISIT_TO_EXISTING_APPLICATION"));self.assertIn("Mobile already exists",preview.rows[0].warnings[0].message);self.assertEqual(resume_import(self.db,self.admin).import_token,preview.import_token)
  other=User(full_name="Other Admin",username="other-admin",email="other-admin@test",password_hash="x",role="Admin");self.db.add(other);self.db.commit()
  with self.assertRaises(PermissionError):commit_import(self.db,other,preview.import_token,[ImportCommitRow(row_number=2,resolved_data=preview.rows[0].data)])
  result=commit_import(self.db,self.admin,preview.import_token,[ImportCommitRow(row_number=2,resolved_data=preview.rows[0].data)]);self.assertEqual((result.imported_rows,result.added_to_existing_applications),(1,1));self.assertEqual(self.db.query(CaseVisit).count(),2)

 def test_downloaded_template_populated_then_previewed(self):
  wb=load_workbook(BytesIO(template_bytes(self.db,self.admin)));ws=wb["Case Import"]
  for index,value in enumerate(self.row(),1):ws.cell(2,index).value=value
  out=BytesIO();wb.save(out);preview=preview_import(self.db,self.admin,out.getvalue(),"case_import_template.xlsx")
  self.assertEqual(preview.summary.total_rows,1);self.assertEqual(preview.rows[0].state,"VALID");self.assertEqual(preview.rows[0].data["los_no"],"LOS-1")

if __name__=="__main__":unittest.main()
